import openpyxl
from openpyxl import *
import constants
import os
import model1
import xmlParse
import mode2

def getColumnData(sheet, column_index):
    column_data = []

    for row in sheet.iter_rows(values_only=True):
        cell = row[column_index]
        cell_value = str(cell) if cell is not None else ""
        if cell_value and cell_value != "变量" and cell_value != "OSD":
            column_data.append(cell_value)

    return column_data

#获取有效列数
def getValidColumnCount(worksheet:Workbook)->int:
    # 获取最大列索引
    max_column = worksheet.max_column
    # 计算有效列数
    valid_column_count = 0
    for column in range(1, max_column + 1):
        cell_value = worksheet.cell(row=1, column=column).value
        if cell_value:
            valid_column_count += 1

    return valid_column_count
def parseXlsx(file_path,callbackMessage):
    worksheet = readExcel(file_path)
    initListConsts(worksheet)
    initMode1(worksheet,callbackMessage)
    

#获取worker
def readExcel(file_path):
    worker = openpyxl.load_workbook(file_path)
    workerSheet = worker.active #取用活动工作铺
    return workerSheet

# 模式1启动
def initMode1(sheet:Workbook,callBack):
    count = getValidColumnCount(sheet)
    for one in range(count):
        if(one>1):
            #过滤第一行再开始输出
            listLanugeStr:list = getColumnData(sheet,one)
            lanHead = constants.getLanguageFilterByTitle(listLanugeStr[0])
            mapN = {}
            if lanHead!="":
                langBean = constants.LanguageBean()
                langBean.languageXmlCode = lanHead
                turelist = listLanugeStr[1:]
                for constValue in constants.listConst:
                    index = constants.listConst.index(constValue)
                    if index <= len(turelist)-1:
                        mapN[constValue] = turelist[index]
                langBean.nameMap = mapN
                constants.listLanguage.append(langBean)
    callBack("模式一读取xslx完成\n")
    model1.genrateXmlFile(callBack)
    
# 模式2启动
def initModel2(oldPath,newPath,callBack):
    callBack("开始解析xml\n")
    mode2.compareXml(oldPath,newPath,callBack)
        


      


#获取变量列
def initListConsts(sheet:Workbook):
    constants.listConst = getColumnData(sheet,0)
    print("基准列表",constants.listConst)
    
